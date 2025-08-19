import streamlit as st
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
import openpyxl
import xlrd
import io
import base64
import os
from typing import List, Tuple, Optional
import numpy as np
from PIL import Image

# Add HEIC/HEIF support
try:
    from pillow_heif import register_heif_opener
    register_heif_opener()
    HEIF_AVAILABLE = True
except ImportError:
    HEIF_AVAILABLE = False

# Try to import QR/CV libraries with fallback
try:
    import cv2
    CV2_AVAILABLE = True
except ImportError:
    CV2_AVAILABLE = False

try:
    from pyzbar.pyzbar import decode as pyzbar_decode
    PYZBAR_AVAILABLE = True
except ImportError:
    PYZBAR_AVAILABLE = False

# QR functionality is available only if both libraries are present
QR_AVAILABLE = CV2_AVAILABLE and PYZBAR_AVAILABLE

# ===================== CONFIGURATION =====================
st.set_page_config(
    page_title="Plant Data Suite",
    page_icon="🌱",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for styling
st.markdown("""
<style>
/* Navigation styling */
.nav-header {
    background: linear-gradient(90deg, #4CAF50, #45a049);
    padding: 1rem;
    border-radius: 10px;
    margin-bottom: 2rem;
    color: white;
    text-align: center;
    font-size: 24px;
    font-weight: bold;
}

/* Section separators */
.section-divider {
    border-top: 3px solid #4CAF50;
    margin: 2rem 0;
    opacity: 0.3;
}

/* Big action buttons */
.big-action-button .stButton > button {
    width: 100% !important;
    height: 120px !important;
    font-size: 28px !important;
    font-weight: bold !important;
    border: none !important;
    border-radius: 15px !important;
    cursor: pointer !important;
    transition: all 0.3s !important;
    padding: 20px 40px !important;
    margin: 15px 0 !important;
}

.big-action-button .stButton > button:hover {
    transform: translateY(-3px) !important;
    box-shadow: 0 8px 16px rgba(0,0,0,0.3) !important;
}

/* Process button styling */
.process-button .stButton > button {
    background-color: #FF6B35 !important;
    color: white !important;
    text-shadow: 1px 1px 2px rgba(0,0,0,0.3) !important;
    box-shadow: 0 4px 8px rgba(255,107,53,0.3) !important;
}

.process-button .stButton > button:hover {
    background-color: #E55A2B !important;
    box-shadow: 0 6px 12px rgba(255,107,53,0.4) !important;
}

/* Download button styling */
.download-button .stButton > button {
    background-color: #4CAF50 !important;
    color: white !important;
    text-shadow: 1px 1px 2px rgba(0,0,0,0.3) !important;
    box-shadow: 0 4px 8px rgba(76,175,80,0.3) !important;
}

.download-button .stButton > button:hover {
    background-color: #45a049 !important;
    box-shadow: 0 6px 12px rgba(76,175,80,0.4) !important;
}

/* Combine button styling */
.combine-button .stButton > button {
    background-color: #2196F3 !important;
    color: white !important;
    text-shadow: 1px 1px 2px rgba(0,0,0,0.3) !important;
    box-shadow: 0 4px 8px rgba(33,150,243,0.3) !important;
}

.combine-button .stButton > button:hover {
    background-color: #1976D2 !important;
    box-shadow: 0 6px 12px rgba(33,150,243,0.4) !important;
}

# QR Reader button styling
.qr-button .stButton > button {
    background-color: #9C27B0 !important;
    color: white !important;
    text-shadow: 1px 1px 2px rgba(0,0,0,0.3) !important;
    box-shadow: 0 4px 8px rgba(156,39,176,0.3) !important;
}

.qr-button .stButton > button:hover {
    background-color: #7B1FA2 !important;
    box-shadow: 0 6px 12px rgba(156,39,176,0.4) !important;
}
.stDownloadButton > button {
    width: 100% !important;
    height: 50px !important;
    font-size: 16px !important;
    font-weight: bold !important;
    background-color: #4CAF50 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    cursor: pointer !important;
    transition: all 0.3s !important;
}

.stDownloadButton > button:hover {
    background-color: #45a049 !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important;
}

/* Function cards */
.function-card {
    border: 2px solid #e0e0e0;
    border-radius: 10px;
    padding: 1rem;
    margin: 1rem 0;
    background-color: #f9f9f9;
}

.function-card.active {
    border-color: #4CAF50;
    background-color: #f0f8f0;
}
</style>
""", unsafe_allow_html=True)

# Template file locations
TEMPLATE_FILE = "z-sheet.xlsx"
LAMP_TEMPLATE = "LAMP-X.xlsx"
QPCR_TEMPLATE = "QPCR-X.xlsx"

def check_template_exists(template_file):
    """Check if a template file exists in the repository."""
    return os.path.exists(template_file)

def load_template_from_file(template_file):
    """Load a template file into a buffer for processing."""
    try:
        with open(template_file, 'rb') as f:
            template_buffer = io.BytesIO(f.read())
            template_buffer.seek(0)
            return template_buffer
    except Exception as e:
        st.error(f"Error loading template file {template_file}: {str(e)}")
        return None

# ===================== PLANT DATA PROCESSOR FUNCTIONS =====================
def standardize_tube(val):
    """Normalize Tube Code to exactly 'TUBE <digits>'."""
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s == "":
        return None

    try:
        f = float(s)
        if f.is_integer():
            return f"TUBE {int(f)}"
    except:
        pass

    nums = re.findall(r'\d+', s)
    if nums:
        longest = max(nums, key=len)
        return f"TUBE {longest}"

    m2 = re.search(r'tube\s*([A-Za-z0-9]+)\s*$', s, flags=re.IGNORECASE)
    if m2:
        token = m2.group(1)
        digits = re.sub(r'\D', '', token)
        return f"TUBE {digits}" if digits else f"TUBE {token}"

    return None

def standardize_clone(val):
    """Keep empty cells empty; convert datetime to YYYY-MM-DD; keep everything else as string."""
    if pd.isna(val) or str(val).strip() == "":
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val)

def make_plant_codes_unique(df):
    """Append (1), (2), (3)... to duplicate Plant Codes to make them unique."""
    counts = {}
    new_codes = []
    for code in df["Plant Code"]:
        if pd.isna(code) or str(code).strip() == "":
            new_codes.append(code)
            continue
        if code not in counts:
            counts[code] = 0
            new_codes.append(code)
        else:
            counts[code] += 1
            new_codes.append(f"{code} ({counts[code]})")
    df["Plant Code"] = new_codes
    return df

def clean_empty(val):
    """Ensure empty/NaN values are written as None (blank cell)."""
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s == "" or s.lower() in ["nan", "none"]:
        return None
    return val

def _finalize_df(df):
    """Select required columns, uniquify Plant Codes, and convert empties to None."""
    required = ["Plant Code", "Tube Code", "Strain", "Clone", "Notes"]
    available = [c for c in required if c in df.columns]
    df = df[available]
    for col in required:
        if col not in df.columns:
            df[col] = None
    df = df[required]
    df = make_plant_codes_unique(df)
    df = df.applymap(clean_empty)
    return df

def clean_old_format(df):
    """Clean the old single-sheet format (CSV or simple XLSX)."""
    if "Number" in df.columns:
        df = df.drop(columns=["Number"])
    if "Tube Code" in df.columns:
        df["Tube Code"] = df["Tube Code"].apply(standardize_tube)
    if "Clone" in df.columns:
        df["Clone"] = df["Clone"].apply(standardize_clone)
    df = _finalize_df(df)
    return df

def clean_new_format(uploaded_file):
    """Clean the new multi-sheet XLSX format by processing only the active sheet."""
    wb = load_workbook(uploaded_file, data_only=True)
    active_sheet = wb.active.title
    
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name=active_sheet)

    normalized_cols = (
        df.columns.str.lower()
        .str.strip()
        .str.replace("*", "", regex=False)
        .str.replace("  ", " ")
    )

    col_map = {}
    for col in normalized_cols:
        if "tube" in col:
            col_map[col] = "Tube Code"
        elif "plant" in col:
            col_map[col] = "Plant Code"
        elif "strain" in col:
            col_map[col] = "Strain"
        elif "clone" in col:
            col_map[col] = "Clone"
        elif "note" in col:
            col_map[col] = "Notes"

    df.columns = [col_map.get(c, c) for c in normalized_cols]
    df = df.loc[:, ~df.columns.duplicated()]

    if "Tube Code" in df.columns:
        df["Tube Code"] = df["Tube Code"].apply(standardize_tube)
    if "Clone" in df.columns:
        df["Clone"] = df["Clone"].apply(standardize_clone)

    df = _finalize_df(df)
    return df

def fill_template(cleaned_df, template_file_buffer):
    """Fill z-sheet template with cleaned data."""
    wb = load_workbook(template_file_buffer)
    ws = wb.active

    column_mapping = {
        "Plant Code": "B",
        "Tube Code": "C",
        "Strain": "E",
        "Clone": "F",
        "Notes": "G"
    }

    for i, row in cleaned_df.iterrows():
        excel_row = i + 2
        for col_name, col_letter in column_mapping.items():
            value = row[col_name]
            ws[f"{col_letter}{excel_row}"] = value if value not in ["", "nan", "NaN"] else None

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

def fill_headwaters_template(cleaned_df, template_file_buffer):
    """Fill z-sheet template with Headwaters data."""
    wb = load_workbook(template_file_buffer)
    ws = wb.active

    # Map DataFrame columns to z-sheet columns
    # DataFrame: Plant Code, Tube Code, , Strain, Clone #, Notes
    # Z-sheet: Plant Code, Tube 1 *, Strain, Clone, Notes
    column_mapping = {
        "Plant Code": "B",      # Plant Code -> Plant Code
        "Tube Code": "C",       # Tube Code -> Tube 1 *
        "Strain": "E",          # Strain -> Strain
        "Clone #": "F",         # Clone # -> Clone
        "Notes": "G"            # Notes -> Notes
    }

    for i, row in cleaned_df.iterrows():
        excel_row = i + 2
        for col_name, col_letter in column_mapping.items():
            value = row[col_name]
            # Convert to string and clean up
            if pd.isna(value) or value == "" or str(value).strip() == "":
                ws[f"{col_letter}{excel_row}"] = None
            else:
                ws[f"{col_letter}{excel_row}"] = str(value).strip()

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

def process_single_file(uploaded_file, filename, template_buffer):
    """Process a single uploaded file."""
    try:
        if filename.endswith(".xlsx"):
            df_clean = clean_new_format(uploaded_file)
        else:
            uploaded_file.seek(0)
            if filename.endswith(".csv"):
                df_raw = pd.read_csv(uploaded_file)
            else:
                df_raw = pd.read_excel(uploaded_file)
            df_clean = clean_old_format(df_raw)

        output_buffer = fill_template(df_clean, template_buffer)
        base_name = filename.rsplit('.', 1)[0]
        output_filename = f"{base_name}_filled.xlsx"
        
        return df_clean, output_buffer, output_filename, None
    except Exception as e:
        return None, None, None, str(e)

# ===================== EXCEL COMBINER FUNCTIONS =====================
def collect_headwaters_data_from_sheets(uploaded_file):
    """Collect data from all sheets in a Bad Client Excel Sheet - adapted from combine2.py."""
    tube_data = []
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Find the header row (look for expected column names)
            header_row = None
            for row_num in range(1, min(20, sheet.max_row + 1)):  # Check first 20 rows
                row_values = [str(cell.value).strip() if cell.value else "" for cell in sheet[row_num]]
                if any("tube" in val.lower() for val in row_values) and any("plant" in val.lower() for val in row_values):
                    header_row = row_num
                    break
            
            if header_row is None:
                continue  # Skip sheets without expected headers
            
            # Get column indices - match original script column structure
            col_indices = {}
            for col_num, cell in enumerate(sheet[header_row], 1):
                cell_value = str(cell.value).strip().lower() if cell.value else ""
                if "tube" in cell_value:
                    col_indices["Tube"] = col_num
                elif "plant" in cell_value:
                    col_indices["Plant Code"] = col_num
                elif "clone" in cell_value:
                    col_indices["Clone #"] = col_num  # Match original script column name
                elif "strain" in cell_value:
                    col_indices["Strain"] = col_num
            
            # Process data rows - extract from column C (Tube) as in original script
            for row_num in range(header_row + 1, sheet.max_row + 1):
                # Use column C (3) for tube data as in original combine2.py
                tube_value = sheet.cell(row=row_num, column=3).value
                
                # Only add rows with tube data and ensure it's not a header
                if tube_value and str(tube_value).strip():
                    tube_str = str(tube_value).strip()
                    # Skip if this looks like a header (contains common header words)
                    header_words = ['tube', 'plant', 'clone', 'strain', 'number', 'code']
                    if not any(word in tube_str.lower() for word in header_words):
                        # Match original script structure: [tube_value, "", "", "", ""]
                        tube_data.append([tube_str, "", "", "", ""])
        
    except Exception as e:
        st.error(f"Error processing file: {str(e)}")
    
    return tube_data

def remove_duplicates(tube_data):
    """Remove duplicate tube entries - match original script logic."""
    seen = set()
    unique_data = []
    
    for row in tube_data:
        tube_id = row[1]  # Tube Code is in position 1 (second column)
        if tube_id not in seen:
            seen.add(tube_id)
            unique_data.append(row)
    
    return unique_data

def normalize_tube_ids(df, column="Tube Code"):
    """Normalize tube IDs for matching."""
    df = df.copy()
    df["_normalized_tube"] = df[column].astype(str).str.strip().str.lower()
    return df

def extract_plant_code(tube_id):
    """Extract plant code from tube ID."""
    match = re.search(r'(\d+)', str(tube_id))
    return match.group(1) if match else ""

def match_and_process(combined_df, reference_df):
    """Match combined data with reference file and process - adapted from excel_cleanup2_final_autofill_plantcode.py."""
    combined_df = normalize_tube_ids(combined_df)
    reference_df = normalize_tube_ids(reference_df)
    
    ref_lookup = reference_df.set_index("_normalized_tube")
    final_rows = []
    
    for _, row in combined_df.iterrows():
        tube_id_norm = row["_normalized_tube"]
        original_tube_id = row["Tube Code"]
        
        if tube_id_norm in ref_lookup.index:
            matched_row = ref_lookup.loc[tube_id_norm]
            if isinstance(matched_row, pd.DataFrame):
                matched_row = matched_row.iloc[0]
            matched_row = matched_row.drop(labels=["_normalized_tube"], errors="ignore")
            
            # Auto-fill plant code if missing
            if pd.isna(matched_row.get("Plant Code", None)) or str(matched_row.get("Plant Code")).strip() == "":
                matched_row["Plant Code"] = extract_plant_code(original_tube_id)
            matched_row["__missing"] = False
            final_rows.append(matched_row)
        else:
            plant_code = extract_plant_code(original_tube_id)
            new_row = {
                "Plant Code": plant_code,
                "Tube Code": original_tube_id,
                "Clone #": "",
                "Strain": "",
                "Notes": "Tube missing from reference Excel sheet",
                "__missing": True
            }
            final_rows.append(pd.Series(new_row))
    
    final_df = pd.DataFrame(final_rows)
    
    # Reorder columns and add empty column
    final_df.insert(2, " ", "")
    final_df = final_df[["Plant Code", "Tube Code", " ", "Strain", "Clone #", "Notes", "__missing"]]
    
    # Sort missing tubes to bottom
    final_df.sort_values(by="__missing", inplace=True)
    final_df.drop(columns=["__missing"], inplace=True)
    
    return final_df
    
# ===================== QR CODE READER FUNCTIONS =====================
def add_white_border(img, pixels=40):
    """Add white border around image for better QR detection."""
    return cv2.copyMakeBorder(
        img, pixels, pixels, pixels, pixels,
        cv2.BORDER_CONSTANT, value=[255, 255, 255]
    )

def try_rotations(img, angles=(15, -15, 30, -30)):
    """Try different rotations to improve QR detection."""
    if not QR_AVAILABLE:
        return None
    
    for angle in angles:
        M = cv2.getRotationMatrix2D((img.shape[1] // 2, img.shape[0] // 2), angle, 1.0)
        rotated = cv2.warpAffine(img, M, (img.shape[1], img.shape[0]), borderValue=(255, 255, 255))
        result = pyzbar_decode(rotated)
        if result:
            return result
    return None

def process_plate_image(uploaded_image, template_buffer, plate_config, scale_factor=1.0):
    """Process a single plate image to extract QR codes."""
    if not QR_AVAILABLE:
        return None, None, "QR code libraries not available. Please install opencv-python and pyzbar."
    
    try:
        # Read image
        image = Image.open(uploaded_image).convert("RGB")
        img = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
        
        # Scale image for higher resolution if requested
        if scale_factor != 1.0:
            height, width = img.shape[:2]
            new_height, new_width = int(height * scale_factor), int(width * scale_factor)
            img = cv2.resize(img, (new_width, new_height), interpolation=cv2.INTER_CUBIC)
        
        # Configuration
        COLS = plate_config.get("cols", 8)
        ROWS = plate_config.get("rows", 12)
        MARGIN = plate_config.get("margin", 12)
        CROP_WIDTH = plate_config.get("crop_width", 2180)
        CROP_HEIGHT = plate_config.get("crop_height", 3940)
        
        # Crop image
        img_h, img_w = img.shape[:2]
        x = max(0, (img_w - CROP_WIDTH) // 2)
        y = max(0, (img_h - CROP_HEIGHT) // 2)
        
        # Ensure we don't exceed image boundaries
        x2 = min(x + CROP_WIDTH, img_w)
        y2 = min(y + CROP_HEIGHT, img_h)
        cropped_img = img[y:y2, x:x2]
        
        # Calculate cell dimensions
        actual_width = cropped_img.shape[1]
        actual_height = cropped_img.shape[0]
        cell_w = actual_width // COLS
        cell_h = actual_height // ROWS
        
        # Load template
        wb = load_workbook(template_buffer)
        ws = wb["samples"] if "samples" in wb.sheetnames else wb.active
        
        # Generate positions
        col_labels = list("ABCDEFGH")
        positions = [f"{col}{row+1}" for row in range(ROWS) for col in col_labels]
        
        # Create debug image
        debug_img = cropped_img.copy()
        results = []
        
        # Process each cell
        for pos in positions:
            row = int(pos[1:]) - 1
            col = 7 - col_labels.index(pos[0])  # A1 top-right
            
            x0 = col * cell_w
            y0 = row * cell_h
            x1 = max(x0 - MARGIN, 0)
            y1 = max(y0 - MARGIN, 0)
            x2 = min(x0 + cell_w + MARGIN, actual_width)
            y2 = min(y0 + cell_h + MARGIN, actual_height)
            
            # Extract cell
            crop = cropped_img[y1:y2, x1:x2]
            if crop.size == 0:
                results.append((pos, ""))
                continue
                
            crop = add_white_border(crop)
            
            # Process for QR detection
            gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
            blur = cv2.GaussianBlur(gray, (3, 3), 0)
            sharp = cv2.addWeighted(gray, 2.0, blur, -1.0, 0)
            _, thresh = cv2.threshold(sharp, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # Try QR detection
            qrs = pyzbar_decode(crop) or pyzbar_decode(gray) or pyzbar_decode(thresh)
            if not qrs:
                qrs = try_rotations(crop) or try_rotations(gray) or try_rotations(thresh)
            
            if qrs:
                qr_data = qrs[0].data.decode("utf-8").strip()
                results.append((pos, qr_data))
                cv2.putText(debug_img, f"{pos}: {qr_data}", (x1 + 5, y1 + 20),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            else:
                results.append((pos, ""))
                cv2.putText(debug_img, f"{pos}: ---", (x1 + 5, y1 + 20),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
            
            # Draw grid
            cv2.rectangle(debug_img, (x1, y1), (x2, y2), (255, 0, 0), 2)
            cv2.putText(debug_img, pos, (x1 + 10, y1 + 50),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
        
        # Sort results and populate Excel
        def well_sort_key(entry):
            col = ord(entry[0][0]) - ord('A')
            row = int(entry[0][1:])
            return (row, col)
        
        results_sorted = sorted(results, key=well_sort_key)
        for idx, (pos, code) in enumerate(results_sorted):
            ws[f"B{idx + 2}"] = pos
            ws[f"C{idx + 2}"] = code
        
        # Save Excel to buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Convert debug image to format for Streamlit
        debug_img_rgb = cv2.cvtColor(debug_img, cv2.COLOR_BGR2RGB)
        debug_img_pil = Image.fromarray(debug_img_rgb)
        
        # Calculate success metrics
        total = len(results_sorted)
        success = sum(1 for _, val in results_sorted if val.strip())
        failed = total - success
        failed_positions = [p for p, v in results_sorted if not v.strip()]
        
        return {
            'excel_buffer': excel_buffer,
            'debug_image': debug_img_pil,
            'results': results_sorted,
            'total': total,
            'success': success,
            'failed': failed,
            'failed_positions': failed_positions
        }, None, None
        
    except Exception as e:
        return None, None, str(e)
def plant_data_processor():
    """Plant Data Processor function."""
    st.markdown('<div class="nav-header">🌱 Plant Data Processor</div>', unsafe_allow_html=True)
    
    # Check if template file exists in repository
    if not check_template_exists(TEMPLATE_FILE):
        st.error(f"❌ Template file '{TEMPLATE_FILE}' not found in repository!")
        st.info(f"Please ensure '{TEMPLATE_FILE}' is in the same directory as this application.")
        return
    
    st.success(f"✅ Template file '{TEMPLATE_FILE}' found in repository!")
    
    # Data files upload
    st.header("📊 Upload Raw Client Sheet")
    uploaded_files = st.file_uploader(
        "Upload your data files (CSV or Excel)",
        type=['csv', 'xlsx'],
        accept_multiple_files=True,
        key="data_files"
    )
    
    if not uploaded_files:
        st.info("Please upload one or more data files to process.")
        return
    
    # Process button
    st.markdown('<div class="big-action-button process-button">', unsafe_allow_html=True)
    process_clicked = st.button("🚀 Process All Files", key="process_data")
    st.markdown('</div>', unsafe_allow_html=True)
    
    if process_clicked:
        results = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, uploaded_file in enumerate(uploaded_files):
            status_text.text(f"Processing {uploaded_file.name}...")
            
            # Load template buffer from repository file
            template_buffer = load_template_from_file(TEMPLATE_FILE)
            if not template_buffer:
                st.error(f"❌ Failed to load template file: {TEMPLATE_FILE}")
                continue
            
            df_clean, output_buffer, output_filename, error = process_single_file(
                uploaded_file, uploaded_file.name, template_buffer
            )
            
            if error:
                st.error(f"❌ Error processing {uploaded_file.name}: {error}")
            else:
                results.append({
                    'original_name': uploaded_file.name,
                    'output_name': output_filename,
                    'data': df_clean,
                    'file_buffer': output_buffer
                })
                st.success(f"✅ Successfully processed {uploaded_file.name}")
            
            progress_bar.progress((i + 1) / len(uploaded_files))
        
        status_text.text("Processing complete!")
        
        if results:
            st.header("📥 Download Results")
            
            for result in results:
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.text(f"📄 {result['output_name']} ({len(result['data'])} rows)")
                with col2:
                    st.download_button(
                        label="📥 Download",
                        data=result['file_buffer'].getvalue(),
                        file_name=result['output_name'],
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        key=f"download_{result['output_name']}"
                    )

def excel_combiner():
    """Headwaters Submission function."""
    st.markdown('<div class="nav-header">🌊 Headwaters Submission</div>', unsafe_allow_html=True)
    
    st.markdown("""
    This tool processes Bad Client Excel Sheets with multiple sheets and creates a standardized z-sheet submission.
    
    **Process:**
    1. Upload Bad Client Excel Sheet (multiple sheets with columns: Tube, Plant Code, Clone Number, Strain)
    2. Upload Reference Excel Sheet (Bad-Client-Excel.xlsx equivalent) for data matching
    3. The tool will extract data from all sheets, remove duplicates, match with reference data, and create a final z-sheet
    """)
    
    # Check if template file exists in repository
    if not check_template_exists(TEMPLATE_FILE):
        st.error(f"❌ Reference file '{TEMPLATE_FILE}' not found in repository!")
        st.info(f"Please ensure '{TEMPLATE_FILE}' is in the same directory as this application.")
        return
    
    st.success(f"✅ Reference file '{TEMPLATE_FILE}' found in repository!")
    
    # Upload Bad Client Excel Sheet
    st.header("📁 Upload Bad Client Excel Sheet")
    bad_client_sheet = st.file_uploader(
        "Upload Bad Client Excel Sheet (.xlsx)",
        type=['xlsx'],
        accept_multiple_files=False,
        key="bad_client_sheet",
        help="Upload Bad Client Excel Sheet with multiple sheets containing Tube, Plant Code, Clone Number, and Strain columns"
    )
    
    # Upload Reference Excel Sheet (Bad-Client-Excel.xlsx equivalent)
    st.header("📋 Upload Reference Excel Sheet")
    reference_sheet = st.file_uploader(
        "Upload Reference Excel Sheet (.xlsx)",
        type=['xlsx'],
        accept_multiple_files=False,
        key="reference_sheet",
        help="Upload reference Excel sheet (Bad-Client-Excel.xlsx) to match against and auto-fill missing data"
    )
    
    if not bad_client_sheet:
        st.info("Please upload a Bad Client Excel Sheet to process.")
        return
    
    if not reference_sheet:
        st.info("Please upload a Reference Excel Sheet to match against.")
        return
    
    # Process button
    st.markdown('<div class="big-action-button combine-button">', unsafe_allow_html=True)
    combine_clicked = st.button("🔄 Combine & Process Files", key="combine_process")
    st.markdown('</div>', unsafe_allow_html=True)
    
    if combine_clicked:
        try:
            # Step 1: Collect data from all sheets
            st.info("🔍 Collecting data from all sheets in Bad Client Excel Sheet...")
            tube_data = collect_headwaters_data_from_sheets(bad_client_sheet)
            st.success(f"✅ Collected {len(tube_data)} total entries from all sheets")
            
            # Step 2: Remove duplicates
            st.info("🧹 Removing duplicates...")
            unique_data = remove_duplicates(tube_data)
            duplicates_removed = len(tube_data) - len(unique_data)
            st.success(f"✅ Removed {duplicates_removed} duplicates, {len(unique_data)} unique entries remain")
            
            # Step 3: Load reference data
            st.info("📋 Loading reference data for matching...")
            reference_df = pd.read_excel(reference_sheet)
            st.success(f"✅ Loaded reference data with {len(reference_df)} entries")
            
            # Step 4: Create combined DataFrame and match with reference
            st.info("🔗 Matching data with reference file...")
            combined_df = pd.DataFrame(unique_data, columns=["Plant Code", "Tube Code", "Strain", "Clone #", "Notes"])
            
            # Match and process data
            final_df = match_and_process(combined_df, reference_df)
            st.success(f"✅ Matching complete! Final dataset has {len(final_df)} entries")
            
            # Step 5: Create final z-sheet
            st.info("📖 Creating final z-sheet...")
            
            # Load template buffer
            template_buffer = load_template_from_file(TEMPLATE_FILE)
            if not template_buffer:
                st.error(f"❌ Failed to load template file: {TEMPLATE_FILE}")
                return
            
            # Fill the z-sheet template with Headwaters data
            output_buffer = fill_headwaters_template(final_df, template_buffer)
            st.success(f"✅ Processing complete! Final z-sheet has {len(final_df)} entries")
            
            # Display results
            st.header("📊 Results Summary")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Entries Collected", len(tube_data))
            with col2:
                st.metric("After Deduplication", len(unique_data))
            with col3:
                st.metric("Reference Entries", len(reference_df))
            with col4:
                st.metric("Final Z-Sheet Entries", len(final_df))
            
            # Download button
            st.header("📥 Download Z-Sheet")
            st.markdown('<div class="big-action-button download-button">', unsafe_allow_html=True)
            st.download_button(
                label="📥 Download Headwaters Z-Sheet",
                data=output_buffer.getvalue(),
                file_name="Headwaters_Submission.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key="download_final"
            )
            st.markdown('</div>', unsafe_allow_html=True)
            
        except Exception as e:
            st.error(f"❌ Error during processing: {str(e)}")

def qr_plate_processor():
    """QR Code Plate Processor function."""
    st.markdown('<div class="nav-header">🔍 QR Code Plate Processor</div>', unsafe_allow_html=True)
    
    # Check if QR libraries are available
    if not QR_AVAILABLE:
        st.error("❌ QR Code processing requires additional libraries!")
        
        # Show specific missing libraries
        missing_libs = []
        if not CV2_AVAILABLE:
            missing_libs.append("opencv-python")
        if not PYZBAR_AVAILABLE:
            missing_libs.append("pyzbar")
        
        st.markdown(f"""
        **Missing libraries:** {', '.join(missing_libs)}
        
        **For local development, install with:**
        ```bash
        pip install {' '.join(missing_libs)}
        ```
        
        **For Streamlit Cloud deployment:**
        
        Add this `packages.txt` file to your repository root:
        ```
        libzbar0
        ```
        
        And this `requirements.txt`:
        ```
        opencv-python-headless
        pyzbar
        ```
        
        Note: OpenCV can be challenging in cloud environments. Use `opencv-python-headless` for better compatibility.
        """)
        
        st.info("💡 **Alternative**: You can use the other two functions (Plant Data Processor and Excel Combiner) which work without these dependencies.")
        return
    
    # Check if template files exist
    lamp_exists = check_template_exists(LAMP_TEMPLATE)
    qpcr_exists = check_template_exists(QPCR_TEMPLATE)
    
    if not lamp_exists and not qpcr_exists:
        st.error(f"❌ Template files not found in repository!")
        st.info(f"Please ensure '{LAMP_TEMPLATE}' and/or '{QPCR_TEMPLATE}' are in the same directory as this application.")
        return
    
    # Show available templates
    st.success("✅ Template files found:")
    col1, col2 = st.columns(2)
    with col1:
        if lamp_exists:
            st.success(f"✅ {LAMP_TEMPLATE}")
        else:
            st.warning(f"⚠️ {LAMP_TEMPLATE} not found")
    with col2:
        if qpcr_exists:
            st.success(f"✅ {QPCR_TEMPLATE}")
        else:
            st.warning(f"⚠️ {QPCR_TEMPLATE} not found")
    
    st.markdown("""
    This tool processes laboratory plate images to extract QR codes and populate Excel templates.
    
    **Process:**
    1. Select template type (LAMP or QPCR)
    2. Upload plate images to process
    3. Configure plate settings
    4. Process images to extract QR codes and generate filled Excel files
    
    **💡 Image Quality Tips:**
    - Use the Image Scale Factor setting (1.5-2.0) for higher resolution processing
    - Annotated images now display at full resolution for better text readability
    - Larger font sizes make QR codes and position labels easier to read
    """)
    
    # Step 1: Template selection
    st.header("🧪 Step 1: Select Template Type")
    template_options = []
    if lamp_exists:
        template_options.append("LAMP")
    if qpcr_exists:
        template_options.append("QPCR")
    
    if not template_options:
        st.error("No valid templates available.")
        return
    
    selected_template = st.radio(
        "Choose template for processing:",
        template_options,
        key="template_choice",
        help=f"Templates are loaded from the repository: {', '.join(template_options)}"
    )
    
    # Step 2: Plate configuration
    st.header("⚙️ Step 2: Plate Configuration")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        cols = st.number_input("Columns", min_value=1, max_value=24, value=8, key="plate_cols")
    with col2:
        rows = st.number_input("Rows", min_value=1, max_value=24, value=12, key="plate_rows")
    with col3:
        margin = st.number_input("Cell Margin", min_value=0, max_value=50, value=12, key="plate_margin")
    with col4:
        pass  # Spacer
    
    # Advanced settings
    with st.expander("🔧 Advanced Settings"):
        col1, col2, col3 = st.columns(3)
        with col1:
            crop_width = st.number_input("Crop Width", min_value=100, max_value=5000, value=2180, key="crop_width")
        with col2:
            crop_height = st.number_input("Crop Height", min_value=100, max_value=5000, value=3940, key="crop_height")
        with col3:
            scale_factor = st.number_input("Image Scale Factor", min_value=0.5, max_value=3.0, value=1.0, step=0.1, 
                                         help="Scale factor for image processing. Higher values (1.5-2.0) may improve QR detection but increase processing time.")
    
    plate_config = {
        "cols": cols,
        "rows": rows,
        "margin": margin,
        "crop_width": crop_width,
        "crop_height": crop_height,
        "scale_factor": scale_factor
    }
    
    # Step 3: Upload images
    st.header("📷 Step 3: Upload Plate Images")
    
    # Show HEIC support status
    if HEIF_AVAILABLE:
        st.success("✅ HEIC/HEIF image format support is available")
    else:
        st.warning("⚠️ HEIC/HEIF support not available - please convert Apple HEIC files to JPG/PNG format first")
        st.info("To add HEIC support, install: `pip install pillow-heif`")
    
    uploaded_images = st.file_uploader(
        "Upload plate images",
        type=['jpg', 'jpeg', 'png', 'heic', 'heif'],
        accept_multiple_files=True,
        key="plate_images",
        help="Upload laboratory plate images for QR code extraction. HEIC files from Apple devices are supported if pillow-heif is installed."
    )
    
    if not uploaded_images:
        st.info("Please upload plate images to process.")
        return
    
    # Process button
    st.markdown('<div class="big-action-button qr-button">', unsafe_allow_html=True)
    process_clicked = st.button("🔍 Process Plate Images", key="process_plates")
    st.markdown('</div>', unsafe_allow_html=True)
    
    if process_clicked:
        # Load the selected template
        template_file = LAMP_TEMPLATE if selected_template == "LAMP" else QPCR_TEMPLATE
        
        results = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, uploaded_image in enumerate(uploaded_images):
            status_text.text(f"Processing {uploaded_image.name}...")
            
            # Load fresh template buffer for each image
            template_buffer = load_template_from_file(template_file)
            if not template_buffer:
                st.error(f"❌ Failed to load template file: {template_file}")
                continue
            
            result, _, error = process_plate_image(uploaded_image, template_buffer, plate_config, plate_config.get("scale_factor", 1.0))
            
            if error:
                st.error(f"❌ Error processing {uploaded_image.name}: {error}")
            elif result:
                base_name = uploaded_image.name.rsplit('.', 1)[0]
                results.append({
                    'original_name': uploaded_image.name,
                    'base_name': base_name,
                    'result': result
                })
                st.success(f"✅ Successfully processed {uploaded_image.name}")
            
            progress_bar.progress((i + 1) / len(uploaded_images))
        
        status_text.text("Processing complete!")
        
        if results:
            st.header("📊 Processing Results")
            
            # Overall statistics
            total_plates = len(results)
            total_wells = sum(result['result']['total'] for result in results)
            total_success = sum(result['result']['success'] for result in results)
            total_failed = total_wells - total_success
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Plates Processed", total_plates)
            with col2:
                st.metric("Total Wells", total_wells)
            with col3:
                st.metric("Successful Reads", total_success)
            with col4:
                st.metric("Failed Reads", total_failed)
            
            # Bulk download section
            st.header("📥 Download Results")
            
            # Create ZIP file for bulk download
            import zipfile
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for result_data in results:
                    result = result_data['result']
                    base_name = result_data['base_name']
                    excel_filename = f"{base_name}_{selected_template}_filled.xlsx"
                    
                    # Add Excel file to ZIP
                    zip_file.writestr(excel_filename, result['excel_buffer'].getvalue())
            
            zip_buffer.seek(0)
            
            # Bulk download button
            st.subheader("📦 Download All Files")
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.markdown('<div class="big-action-button download-button">', unsafe_allow_html=True)
                st.download_button(
                    label="📦 Download All Files (ZIP)",
                    data=zip_buffer.getvalue(),
                    file_name=f"QR_Processing_Results_{selected_template}.zip",
                    mime="application/zip",
                    key="bulk_download_qr"
                )
                st.markdown('</div>', unsafe_allow_html=True)
            
            st.markdown("---")
            st.subheader("📄 Individual Files")
            
            # Individual results
            for idx, result_data in enumerate(results):
                result = result_data['result']
                
                st.subheader(f"📄 {result_data['original_name']}")
                
                col1, col2 = st.columns([2, 1])
                
                with col1:
                    # Statistics
                    success_rate = (result['success'] / result['total']) * 100 if result['total'] > 0 else 0
                    st.write(f"**Success Rate:** {success_rate:.1f}% ({result['success']}/{result['total']})")
                    st.write(f"**Template Used:** {selected_template}")
                    
                    if result['failed'] > 0:
                        st.write(f"**Failed Wells:** {', '.join(result['failed_positions'][:10])}")
                        if len(result['failed_positions']) > 10:
                            st.write(f"... and {len(result['failed_positions']) - 10} more")
                
                with col2:
                    # Download Excel
                    excel_filename = f"{result_data['base_name']}_{selected_template}_filled.xlsx"
                    st.download_button(
                        label="📥 Download Excel",
                        data=result['excel_buffer'].getvalue(),
                        file_name=excel_filename,
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        key=f"excel_{idx}_{result_data['base_name']}"
                    )
                
                # Show debug image
                with st.expander(f"🔍 View Annotated Image - {result_data['original_name']}"):
                    st.image(result['debug_image'], caption=f"Processed plate with QR detection results")
                
                st.markdown("---")

def function_3():
    """QR Code Plate Processor function (renamed from placeholder)."""
    return qr_plate_processor()

def main():
    """Main application function."""
    # Sidebar navigation
    st.sidebar.title("🌱 Plant Data Suite")
    st.sidebar.markdown("---")
    
    # Navigation options
    app_mode = st.sidebar.radio(
        "Choose Function:",
        [
            "🌱 Plant Data Processor", 
            "🌊 Headwaters Submission",
            "🔍 QR Code Plate Processor"
        ],
        key="main_nav"
    )
    
    # Function descriptions in sidebar
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📖 Function Descriptions")
    
    if "Plant Data Processor" in app_mode:
        st.sidebar.markdown("""
        **🌱 Plant Data Processor**
        - Standardize plant data formats
        - Fill template spreadsheets
        - Process multiple files at once
        """)
    elif "Headwaters Submission" in app_mode:
        st.sidebar.markdown("""
        **🌊 Headwaters Submission**
        - Process Bad Client Excel Sheets with multiple sheets
        - Extract data from columns: Tube, Plant Code, Clone Number, Strain
        - Match with reference data and auto-fill missing information
        - Create standardized z-sheet submission
        """)
    else:
        st.sidebar.markdown("""
        **🔍 QR Code Plate Processor**
        - Process laboratory plate images
        - Extract QR codes automatically
        - Generate filled Excel templates
        """)
    
    # Route to appropriate function
    if "Plant Data Processor" in app_mode:
        plant_data_processor()
    elif "Headwaters Submission" in app_mode:
        excel_combiner()
    elif "QR Code Plate Processor" in app_mode:
        qr_plate_processor()

if __name__ == "__main__":
    main()
