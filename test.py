import streamlit as st
import pandas as pd
import re
from datetime import datetime, date
from openpyxl import load_workbook
import openpyxl
import xlrd
import io
import base64
import os
from typing import List, Tuple, Optional
import numpy as np
from PIL import Image
import json
import time
import hashlib
import tempfile

# Add HEIC/HEIF support
try:
    from pillow_heif import register_heif_opener
    register_heif_opener()
    HEIF_AVAILABLE = True
except ImportError:
    HEIF_AVAILABLE = False

# Try to import QR/CV libraries with fallback
try:
    import cv2
    CV2_AVAILABLE = True
except ImportError:
    CV2_AVAILABLE = False

try:
    from pyzbar.pyzbar import decode as pyzbar_decode
    PYZBAR_AVAILABLE = True
except ImportError:
    PYZBAR_AVAILABLE = False

# QR functionality is available only if both libraries are present
QR_AVAILABLE = CV2_AVAILABLE and PYZBAR_AVAILABLE

# ===================== CONFIGURATION =====================
st.set_page_config(
    page_title="The Riaz Machine",
    page_icon="🌱",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for styling
st.markdown("""
<style>
/* Navigation styling */
.nav-header {
    background: rgba(0, 0, 0, 0.7);
    padding: 1.5rem;
    margin-bottom: 2rem;
    color: #ffffff;
    text-align: center;
    font-size: 28px;
    font-weight: bold;
    letter-spacing: 0.5px;
}

/* Section separators */
.section-divider {
    border-top: 3px solid #4CAF50;
    margin: 2rem 0;
    opacity: 0.3;
}

/* Big action buttons */
.big-action-button .stButton > button {
    width: 100% !important;
    height: 120px !important;
    font-size: 28px !important;
    font-weight: bold !important;
    border: none !important;
    border-radius: 15px !important;
    cursor: pointer !important;
    transition: all 0.3s !important;
    padding: 20px 40px !important;
    margin: 15px 0 !important;
}

.big-action-button .stButton > button:hover {
    transform: translateY(-3px) !important;
    box-shadow: 0 8px 16px rgba(0,0,0,0.3) !important;
}

/* Process button styling */
.process-button .stButton > button {
    background-color: #FF6B35 !important;
    color: white !important;
    text-shadow: 1px 1px 2px rgba(0,0,0,0.3) !important;
    box-shadow: 0 4px 8px rgba(255,107,53,0.3) !important;
}

.process-button .stButton > button:hover {
    background-color: #E55A2B !important;
    box-shadow: 0 6px 12px rgba(255,107,53,0.4) !important;
}

/* Download button styling */
.download-button .stButton > button {
    background-color: #4CAF50 !important;
    color: white !important;
    text-shadow: 1px 1px 2px rgba(0,0,0,0.3) !important;
    box-shadow: 0 4px 8px rgba(76,175,80,0.3) !important;
}

.download-button .stButton > button:hover {
    background-color: #45a049 !important;
    box-shadow: 0 6px 12px rgba(76,175,80,0.4) !important;
}

/* Combine button styling */
.combine-button .stButton > button {
    background-color: #2196F3 !important;
    color: white !important;
    text-shadow: 1px 1px 2px rgba(0,0,0,0.3) !important;
    box-shadow: 0 4px 8px rgba(33,150,243,0.3) !important;
}

.combine-button .stButton > button:hover {
    background-color: #1976D2 !important;
    box-shadow: 0 6px 12px rgba(33,150,243,0.4) !important;
}

/* QR Reader button styling */
.qr-button .stButton > button {
    background-color: #9C27B0 !important;
    color: white !important;
    text-shadow: 1px 1px 2px rgba(0,0,0,0.3) !important;
    box-shadow: 0 4px 8px rgba(156,39,176,0.3) !important;
}

.qr-button .stButton > button:hover {
    background-color: #7B1FA2 !important;
    box-shadow: 0 6px 12px rgba(156,39,176,0.4) !important;
}

/* Share button styling */
.share-button .stButton > button {
    background-color: #FF9800 !important;
    color: white !important;
    text-shadow: 1px 1px 2px rgba(0,0,0,0.3) !important;
    box-shadow: 0 4px 8px rgba(255,152,0,0.3) !important;
}

.share-button .stButton > button:hover {
    background-color: #F57C00 !important;
    box-shadow: 0 6px 12px rgba(255,152,0,0.4) !important;
}

.stDownloadButton > button {
    width: 100% !important;
    height: 50px !important;
    font-size: 16px !important;
    font-weight: bold !important;
    background-color: #4CAF50 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    cursor: pointer !important;
    transition: all 0.3s !important;
}

.stDownloadButton > button:hover {
    background-color: #45a049 !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important;
}

/* Function cards */
.function-card {
    border: 2px solid #e0e0e0;
    border-radius: 10px;
    padding: 1rem;
    margin: 1rem 0;
    background-color: #f9f9f9;
}

.function-card.active {
    border-color: #4CAF50;
    background-color: #f0f8f0;
}

/* Share code styling */
.share-code {
    background-color: #f0f0f0;
    border: 2px solid #4CAF50;
    border-radius: 10px;
    padding: 20px;
    text-align: center;
    font-size: 24px;
    font-weight: bold;
    color: #2E7D32;
    margin: 20px 0;
}
</style>
""", unsafe_allow_html=True)

# Template file locations
TEMPLATE_FILE = "z-sheet.xlsx"
LAMP_TEMPLATE = "LAMP-X.xlsx"
QPCR_TEMPLATE = "QPCR-X.xlsx"

# File sharing system
TEMP_DIR = tempfile.gettempdir()
SHARED_FILES_DIR = os.path.join(TEMP_DIR, "riaz_machine_shared")

def init_shared_files_system():
    """Initialize the shared files directory."""
    if not os.path.exists(SHARED_FILES_DIR):
        os.makedirs(SHARED_FILES_DIR)

def generate_share_code():
    """Generate a unique 6-digit share code."""
    timestamp = str(int(time.time()))
    random_str = str(hash(timestamp + str(np.random.random())))
    return hashlib.md5(random_str.encode()).hexdigest()[:6].upper()

def save_shared_files(files, metadata=None):
    """Save files with a unique share code and return the code."""
    init_shared_files_system()
    share_code = generate_share_code()
    share_dir = os.path.join(SHARED_FILES_DIR, share_code)
    os.makedirs(share_dir, exist_ok=True)
    
    # Save metadata
    if metadata is None:
        metadata = {"upload_time": time.time(), "file_count": len(files)}
    
    # Save files
    saved_files = []
    for file in files:
        file_path = os.path.join(share_dir, file.name)
        with open(file_path, "wb") as f:
            f.write(file.getvalue())
        saved_files.append({
            "name": file.name,
            "size": len(file.getvalue()),
            "path": file_path
        })
    
    # Save metadata
    metadata["files"] = saved_files
    metadata_path = os.path.join(share_dir, "metadata.json")
    with open(metadata_path, "w") as f:
        json.dump(metadata, f)
    
    return share_code

def load_shared_files(share_code):
    """Load files using a share code."""
    init_shared_files_system()
    share_dir = os.path.join(SHARED_FILES_DIR, share_code.upper())
    metadata_path = os.path.join(share_dir, "metadata.json")
    
    if not os.path.exists(metadata_path):
        return None
    
    try:
        with open(metadata_path, "r") as f:
            metadata = json.load(f)
        
        files = []
        for file_info in metadata.get("files", []):
            if os.path.exists(file_info["path"]):
                with open(file_info["path"], "rb") as f:
                    file_content = io.BytesIO(f.read())
                    file_content.name = file_info["name"]
                    files.append(file_content)
        
        return {"files": files, "metadata": metadata}
    except Exception as e:
        st.error(f"Error loading shared files: {str(e)}")
        return None

def cleanup_old_shares(max_age_hours=24):
    """Clean up shared files older than max_age_hours."""
    init_shared_files_system()
    current_time = time.time()
    
    for share_code in os.listdir(SHARED_FILES_DIR):
        share_dir = os.path.join(SHARED_FILES_DIR, share_code)
        metadata_path = os.path.join(share_dir, "metadata.json")
        
        if os.path.exists(metadata_path):
            try:
                with open(metadata_path, "r") as f:
                    metadata = json.load(f)
                
                upload_time = metadata.get("upload_time", 0)
                age_hours = (current_time - upload_time) / 3600
                
                if age_hours > max_age_hours:
                    import shutil
                    shutil.rmtree(share_dir)
            except Exception:
                pass

def check_template_exists(template_file):
    """Check if a template file exists in the repository."""
    return os.path.exists(template_file)

def load_template_from_file(template_file):
    """Load a template file into a buffer for processing."""
    try:
        with open(template_file, 'rb') as f:
            template_buffer = io.BytesIO(f.read())
            template_buffer.seek(0)
            return template_buffer
    except Exception as e:
        st.error(f"Error loading template file {template_file}: {str(e)}")
        return None

# ===================== NEW OPTIMIZED PLANT DATA PROCESSOR FUNCTIONS =====================
def standardize_tube(val):
    """Normalize Tube Code to exactly 'TUBE <digits>'. Empty/NaN/blank -> None."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # 1) numeric-looking inputs first
    try:
        f = float(s)
        if f.is_integer():
            return f"TUBE {int(f)}"
    except Exception:
        pass
    # 2) longest digit sequence
    nums = re.findall(r'\d+', s)
    if nums:
        return f"TUBE {max(nums, key=len)}"
    # 3) fallback: 'tube <token>'
    m2 = re.search(r'tube\s*([A-Za-z0-9]+)\s*$', s, flags=re.IGNORECASE)
    if m2:
        token = m2.group(1)
        digits = re.sub(r'\D', '', token)
        return f"TUBE {digits}" if digits else f"TUBE {token}"
    return None

def _parse_date_like_to_yyyy_mm_dd(s: str):
    """Parse any date-like string to YYYY-MM-DD (or None)."""
    s = s.strip()
    if not s:
        return None
    s = re.sub(r"\s+\d{2}:\d{2}:\d{2}$", "", s)  # strip trailing time if present
    try:
        ts = pd.to_datetime(s, errors="raise", infer_datetime_format=True)
        return ts.date().isoformat()
    except Exception:
        pass
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None

def standardize_clone(val):
    """
    Convert anything date-like to 'YYYY-MM-DD'.
    - datetime/date -> YYYY-MM-DD
    - strings like '5/26/2025' or '2025-05-26 00:00:00' -> '2025-05-26'
    - empty -> None
    - non-date text -> returned as-is (trimmed)
    """
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if s == "":
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    parsed = _parse_date_like_to_yyyy_mm_dd(s)
    return parsed if parsed is not None else s

def make_plant_codes_unique_vectorized(df: pd.DataFrame) -> pd.DataFrame:
    """Vectorized suffix: (1), (2), ... for duplicate Plant Codes (skip if none)."""
    df = df.copy()
    pc = df["Plant Code"]
    mask_nonempty = pc.notna() & (pc.astype(str).str.strip() != "")
    if not pc[mask_nonempty].duplicated().any():
        return df  # fast exit: no duplicates

    counts = (
        pc[mask_nonempty]
        .groupby(pc[mask_nonempty])
        .cumcount()
    )
    suffix = counts.where(counts == 0, "(" + counts.astype(str) + ")")
    new_codes = pc[mask_nonempty].astype(str).str.strip()
    new_codes = new_codes.where(suffix.isna() | (suffix == 0),
                                new_codes + " " + suffix.astype(str))
    df.loc[mask_nonempty, "Plant Code"] = new_codes
    return df

def vector_clean_empty(df: pd.DataFrame) -> pd.DataFrame:
    """Vectorized empty/NaN -> None for the entire DF."""
    df = df.replace(r"^\s*$", pd.NA, regex=True)
    df = df.replace({"nan": pd.NA, "NaN": pd.NA, "None": pd.NA, "none": pd.NA})
    return df.where(pd.notna(df), None)

def _finalize_df(df: pd.DataFrame, drop_rows_without_tube=False) -> pd.DataFrame:
    """
    Ensure required columns exist, unique Plant Codes, blanks->None.
    Required: Plant Code, Tube Code, Strain, Clone, Notes
    """
    required = ["Plant Code", "Tube Code", "Strain", "Clone", "Notes"]
    present = [c for c in required if c in df.columns]
    df = df[present].copy()
    for col in required:
        if col not in df.columns:
            df[col] = None
    df = df[required]

    if drop_rows_without_tube:
        df = df[df["Tube Code"].notna()]

    df = make_plant_codes_unique_vectorized(df)
    df = vector_clean_empty(df)
    return df

def _normalize_columns_fuzzy(df: pd.DataFrame) -> pd.DataFrame:
    """Fuzzy auto-map headers to our standard names."""
    norm = (
        pd.Index(df.columns).astype(str).str.lower()
        .str.strip()
        .str.replace("*", "", regex=False)
        .str.replace("  ", " ")
    )
    col_map = {}
    for col in norm:
        if   "tube"   in col: col_map[col] = "Tube Code"
        elif "plant"  in col: col_map[col] = "Plant Code"
        elif "strain" in col or "variety" in col or "cultivar" in col:
            col_map[col] = "Strain"
        elif "clone"  in col:
            col_map[col] = "Clone"
        elif "note"   in col or "remark" in col:
            col_map[col] = "Notes"
    df = df.copy()
    df.columns = [col_map.get(c, c) for c in norm]
    df = df.loc[:, ~df.columns.duplicated()]
    return df

def get_active_sheet_name_from_buffer(uploaded_file) -> str:
    """Get active sheet name from uploaded file buffer."""
    wb = load_workbook(uploaded_file, data_only=True, read_only=True)
    return wb.active.title

def is_special_client_by_header(uploaded_file, active_sheet: str) -> bool:
    """
    Detect special client by checking if D1 says 'clone number' (case-insensitive).
    """
    uploaded_file.seek(0)
    cols = pd.read_excel(uploaded_file, sheet_name=active_sheet, nrows=0).columns
    if len(cols) >= 4:
        d1 = str(cols[3]).strip().lower()
        return d1 == "clone number"
    return False

# ===================== CLEANERS =====================
def clean_old_format(df: pd.DataFrame) -> pd.DataFrame:
    """Clean old single-sheet format (CSV/simple XLSX with exact headers)."""
    if "Number" in df.columns:
        df = df.drop(columns=["Number"])
    if "Tube Code" in df.columns:
        df["Tube Code"] = df["Tube Code"].apply(standardize_tube)
    if "Clone" in df.columns:
        df["Clone"] = df["Clone"].apply(standardize_clone)
    return _finalize_df(df)

def clean_new_format(uploaded_file, active_sheet: str) -> pd.DataFrame:
    """Clean multi-sheet XLSX by processing only the active sheet; fuzzy header mapping."""
    uploaded_file.seek(0)
    df = pd.read_excel(
        uploaded_file,
        sheet_name=active_sheet,
        dtype=str,
        keep_default_na=False
    )
    df = _normalize_columns_fuzzy(df)
    if "Tube Code" in df.columns:
        df["Tube Code"] = df["Tube Code"].apply(standardize_tube)
    if "Clone" in df.columns:
        df["Clone"] = df["Clone"].apply(standardize_clone)
    return _finalize_df(df)

def clean_special_client_all_sheets(uploaded_file) -> pd.DataFrame:
    """
    SPECIAL CASE: D1 == 'clone number' => combine ALL sheets.
    - Fuzzy-map each sheet
    - Standardize Tube/Clone
    - Drop rows with empty Tube Code
    - Concatenate and finalize
    """
    uploaded_file.seek(0)
    xf = pd.ExcelFile(uploaded_file)
    frames = []
    for name in xf.sheet_names:
        df = xf.parse(name, dtype=str, keep_default_na=False)
        if df is None or df.empty:
            continue
        df = _normalize_columns_fuzzy(df)
        if "Tube Code" in df.columns:
            df["Tube Code"] = df["Tube Code"].apply(standardize_tube)
        if "Clone" in df.columns:
            df["Clone"] = df["Clone"].apply(standardize_clone)
        df = _finalize_df(df, drop_rows_without_tube=True)
        if not df.empty:
            frames.append(df)
    if frames:
        combined = pd.concat(frames, ignore_index=True)
        combined = make_plant_codes_unique_vectorized(combined)
        return vector_clean_empty(combined)
    return _finalize_df(pd.DataFrame(columns=["Plant Code","Tube Code","Strain","Clone","Notes"]),
                        drop_rows_without_tube=True)

def fill_template(cleaned_df: pd.DataFrame, template_buffer):
    """
    Fill z-sheet template with cleaned data, writing None for empty cells.
    Clone values are already 'YYYY-MM-DD' strings (or None), so no 00:00:00.
    """
    wb = load_workbook(template_buffer)
    ws = wb.active

    for idx, row in enumerate(cleaned_df.itertuples(index=False), start=2):
        plant, tube, strain, clone, notes = row
        ws.cell(row=idx, column=2, value=None if plant in ("", "nan", "NaN") else plant)   # B
        ws.cell(row=idx, column=3, value=None if tube  in ("", "nan", "NaN") else tube)    # C
        ws.cell(row=idx, column=5, value=None if strain in ("", "nan", "NaN") else strain) # E
        ws.cell(row=idx, column=6, value=None if clone in ("", "nan", "NaN") else clone)   # F
        ws.cell(row=idx, column=7, value=None if notes in ("", "nan", "NaN") else notes)   # G

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

def process_single_file(uploaded_file, filename, template_buffer):
    """Process a single uploaded file using the new optimized logic."""
    try:
        if filename.endswith(".xlsx"):
            # Get active sheet name
            uploaded_file.seek(0)
            active_sheet = get_active_sheet_name_from_buffer(uploaded_file)
            
            # Check if it's a special client
            uploaded_file.seek(0)
            if is_special_client_by_header(uploaded_file, active_sheet):
                st.info(f"🔍 Detected special client format in {filename} - processing all sheets")
                df_clean = clean_special_client_all_sheets(uploaded_file)
            else:
                df_clean = clean_new_format(uploaded_file, active_sheet)
        else:
            # CSV or simple Excel with exact headers
            uploaded_file.seek(0)
            if filename.endswith(".csv"):
                df_raw = pd.read_csv(uploaded_file, dtype=str, keep_default_na=False)
            else:
                df_raw = pd.read_excel(uploaded_file, dtype=str, keep_default_na=False)
            df_clean = clean_old_format(df_raw)

        output_buffer = fill_template(df_clean, template_buffer)
        base_name = filename.rsplit('.', 1)[0]
        output_filename = f"{base_name}_filled.xlsx"
        
        return df_clean, output_buffer, output_filename, None
    except Exception as e:
        return None, None, None, str(e)

# ===================== QR CODE READER FUNCTIONS =====================
def detect_image_orientation(img):
    """Detect if image should be rotated for portrait processing."""
    height, width = img.shape[:2]
    # If width > height, it's landscape and should be rotated to portrait
    if width > height:
        # Rotate 90 degrees counterclockwise to make it portrait
        img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
    return img

def generate_safe_filename(original_name, existing_names=None):
    """Generate a safe filename and avoid duplicates."""
    if existing_names is None:
        existing_names = set()
    
    # Clean the filename
    base_name = os.path.splitext(original_name)[0]
    extension = os.path.splitext(original_name)[1]
    
    # Remove problematic characters
    safe_name = re.sub(r'[<>:"/\\|?*]', '_', base_name)
    safe_name = re.sub(r'\s+', '_', safe_name)
    
    # Handle duplicates
    counter = 1
    final_name = f"{safe_name}{extension}"
    while final_name in existing_names:
        final_name = f"{safe_name}_{counter}{extension}"
        counter += 1
    
    return final_name

def add_white_border(img, pixels=40):
    """Add white border around image for better QR detection."""
    return cv2.copyMakeBorder(
        img, pixels, pixels, pixels, pixels,
        cv2.BORDER_CONSTANT, value=[255, 255, 255]
    )

def try_rotations(img, angles=(15, -15, 30, -30)):
    """Try different rotations to improve QR detection."""
    if not QR_AVAILABLE:
        return None
    
    for angle in angles:
        M = cv2.getRotationMatrix2D((img.shape[1] // 2, img.shape[0] // 2), angle, 1.0)
        rotated = cv2.warpAffine(img, M, (img.shape[1], img.shape[0]), borderValue=(255, 255, 255))
        result = pyzbar_decode(rotated)
        if result:
            return result
    return None

def process_plate_image(uploaded_image, template_buffer, plate_config, scale_factor=1.0):
    """Process a single plate image to extract QR codes with proper orientation handling."""
    if not QR_AVAILABLE:
        return None, None, "QR code libraries not available. Please install opencv-python and pyzbar."
    
    try:
        # Read image
        image = Image.open(uploaded_image).convert("RGB")
        img = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
        
        # FIX: Detect and correct orientation for portrait processing
        img = detect_image_orientation(img)
        
        # Scale image for higher resolution if requested
        if scale_factor != 1.0:
            height, width = img.shape[:2]
            new_height, new_width = int(height * scale_factor), int(width * scale_factor)
            img = cv2.resize(img, (new_width, new_height), interpolation=cv2.INTER_CUBIC)
        
        # Configuration
        COLS = plate_config.get("cols", 8)
        ROWS = plate_config.get("rows", 12)
        MARGIN = plate_config.get("margin", 12)
        CROP_WIDTH = plate_config.get("crop_width", 2180)
        CROP_HEIGHT = plate_config.get("crop_height", 3940)
        
        # Crop image
        img_h, img_w = img.shape[:2]
        x = max(0, (img_w - CROP_WIDTH) // 2)
        y = max(0, (img_h - CROP_HEIGHT) // 2)
        
        # Ensure we don't exceed image boundaries
        x2 = min(x + CROP_WIDTH, img_w)
        y2 = min(y + CROP_HEIGHT, img_h)
        cropped_img = img[y:y2, x:x2]
        
        # Calculate cell dimensions
        actual_width = cropped_img.shape[1]
        actual_height = cropped_img.shape[0]
        cell_w = actual_width // COLS
        cell_h = actual_height // ROWS
        
        # Load template
        wb = load_workbook(template_buffer)
        ws = wb["samples"] if "samples" in wb.sheetnames else wb.active
        
        # Generate positions
        col_labels = list("ABCDEFGH")
        positions = [f"{col}{row+1}" for row in range(ROWS) for col in col_labels]
        
        # Create debug image
        debug_img = cropped_img.copy()
        results = []
        
        # Process each cell
        for pos in positions:
            row = int(pos[1:]) - 1
            col = 7 - col_labels.index(pos[0])  # A1 top-right
            
            x0 = col * cell_w
            y0 = row * cell_h
            x1 = max(x0 - MARGIN, 0)
            y1 = max(y0 - MARGIN, 0)
            x2 = min(x0 + cell_w + MARGIN, actual_width)
            y2 = min(y0 + cell_h + MARGIN, actual_height)
            
            # Extract cell
            crop = cropped_img[y1:y2, x1:x2]
            if crop.size == 0:
                results.append((pos, ""))
                continue
                
            crop = add_white_border(crop)
            
            # Process for QR detection
            gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
            blur = cv2.GaussianBlur(gray, (3, 3), 0)
            sharp = cv2.addWeighted(gray, 2.0, blur, -1.0, 0)
            _, thresh = cv2.threshold(sharp, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # Try QR detection
            qrs = pyzbar_decode(crop) or pyzbar_decode(gray) or pyzbar_decode(thresh)
            if not qrs:
                qrs = try_rotations(crop) or try_rotations(gray) or try_rotations(thresh)
            
            if qrs:
                qr_data = qrs[0].data.decode("utf-8").strip()
                results.append((pos, qr_data))
                cv2.putText(debug_img, f"{pos}: {qr_data}", (x1 + 5, y1 + 20),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            else:
                results.append((pos, ""))
                cv2.putText(debug_img, f"{pos}: ---", (x1 + 5, y1 + 20),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
            
            # Draw grid
            cv2.rectangle(debug_img, (x1, y1), (x2, y2), (255, 0, 0), 2)
            cv2.putText(debug_img, pos, (x1 + 10, y1 + 50),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
        
        # Sort results and populate Excel
        def well_sort_key(entry):
            col = ord(entry[0][0]) - ord('A')
            row = int(entry[0][1:])
            return (row, col)
        
        results_sorted = sorted(results, key=well_sort_key)
        for idx, (pos, code) in enumerate(results_sorted):
            ws[f"B{idx + 2}"] = pos
            ws[f"C{idx + 2}"] = code
        
        # Save Excel to buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Convert debug image to format for Streamlit
        debug_img_rgb = cv2.cvtColor(debug_img, cv2.COLOR_BGR2RGB)
        debug_img_pil = Image.fromarray(debug_img_rgb)
        
        # Calculate success metrics
        total = len(results_sorted)
        success = sum(1 for _, val in results_sorted if val.strip())
        failed = total - success
        failed_positions = [p for p, v in results_sorted if not v.strip()]
        
        return {
            'excel_buffer': excel_buffer,
            'debug_image': debug_img_pil,
            'results': results_sorted,
            'total': total,
            'success': success,
            'failed': failed,
            'failed_positions': failed_positions
        }, None, None
        
    except Exception as e:
        return None, None, str(e)

# ===================== MOBILE FILE SHARING FUNCTION (REMOVED - NOW INTEGRATED INTO QR PROCESSOR) =====================
# This function is no longer needed as sharing is integrated into qr_plate_processor()

# ===================== PLANT DATA PROCESSOR FUNCTIONS =====================
def unified_processor():
    """Unified processor function that handles both plant data and headwaters processing."""
    st.markdown('<div class="nav-header">🔄 Unified Plant Data Processor</div>', unsafe_allow_html=True)
    
    # Check if template file exists in repository
    if not check_template_exists(TEMPLATE_FILE):
        st.error(f"❌ Template file '{TEMPLATE_FILE}' not found in repository!")
        st.info(f"Please ensure '{TEMPLATE_FILE}' is in the same directory as this application.")
        return
    
    st.success(f"✅ Template file '{TEMPLATE_FILE}' found in repository!")
    
    # Data files upload - simplified and prominent
    st.header("📊 Upload Data Files")
    
    uploaded_files = st.file_uploader(
        "Upload your data files (CSV or Excel)",
        type=['csv', 'xlsx'],
        accept_multiple_files=True,
        key="unified_data_files"
    )
    
    # Info section moved to expander
    with st.expander("ℹ️ About This Tool", expanded=False):
        st.markdown("""
        **All-in-One Processing Solution:**
        - 🌱 **Standard Plant Data**: Process individual files with exact column headers
        - 🔍 **Smart Detection**: Automatically detects and handles special client formats
        - 🌊 **Multi-Sheet Processing**: Combines data from all sheets when appropriate
        - ⚡ **High Performance**: Optimized processing with intelligent format detection
        
        This unified tool replaces both the Plant Data Processor and Headwaters Submission functions with enhanced capabilities.
        
        **Processing Logic:**
        - Files with "Clone Number" in column D will have **all sheets processed and combined**
        - Other Excel files will process only the **active sheet**
        - CSV files will be processed as **standard format**
        - Fuzzy column matching works for various naming conventions
        """)
    
    if not uploaded_files:
        st.info("Please upload one or more data files to process.")
        return
    
    # Process button
    st.markdown('<div class="big-action-button process-button">', unsafe_allow_html=True)
    process_clicked = st.button("🚀 Process All Files", key="unified_process")
    st.markdown('</div>', unsafe_allow_html=True)
    
    if process_clicked:
        results = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, uploaded_file in enumerate(uploaded_files):
            status_text.text(f"Processing {uploaded_file.name}...")
            
            # Load template buffer from repository file
            template_buffer = load_template_from_file(TEMPLATE_FILE)
            if not template_buffer:
                st.error(f"❌ Failed to load template file: {TEMPLATE_FILE}")
                continue
            
            df_clean, output_buffer, output_filename, error = process_single_file(
                uploaded_file, uploaded_file.name, template_buffer
            )
            
            if error:
                st.error(f"❌ Error processing {uploaded_file.name}: {error}")
            else:
                results.append({
                    'original_name': uploaded_file.name,
                    'output_name': output_filename,
                    'data': df_clean,
                    'file_buffer': output_buffer
                })
                st.success(f"✅ Successfully processed {uploaded_file.name}")
            
            progress_bar.progress((i + 1) / len(uploaded_files))
        
        status_text.text("Processing complete!")
        
        if results:
            st.header("📊 Processing Results")
            
            # Show summary statistics
            total_files = len(results)
            total_rows = sum(len(result['data']) for result in results)
            
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Files Processed", total_files)
            with col2:
                st.metric("Total Rows Processed", total_rows)
            
            # Show sample of processed data
            st.subheader("📋 Sample of Processed Data")
            sample_df = results[0]['data'].head(10)
            st.dataframe(sample_df)
            
            st.header("📥 Download Results")
            
            # Create ZIP file for bulk download
            import zipfile
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for result in results:
                    zip_file.writestr(result['output_name'], result['file_buffer'].getvalue())
            
            zip_buffer.seek(0)
            
            # Bulk download button
            if len(results) > 1:
                st.subheader("📦 Download All Files")
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                    st.markdown('<div class="big-action-button download-button">', unsafe_allow_html=True)
                    st.download_button(
                        label="📦 Download All Files (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name="Unified_Processing_Results.zip",
                        mime="application/zip",
                        key="bulk_download_unified"
                    )
                    st.markdown('</div>', unsafe_allow_html=True)
                
                st.markdown("---")
                st.subheader("📄 Individual Files")
            
            # Individual file downloads
            for result in results:
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.text(f"📄 {result['output_name']} ({len(result['data'])} rows)")
                with col2:
                    st.download_button(
                        label="📥 Download",
                        data=result['file_buffer'].getvalue(),
                        file_name=result['output_name'],
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        key=f"download_unified_{result['output_name']}"
                    )

def qr_plate_processor():
    """QR Code Plate Processor function with integrated file sharing and improved UI."""
    st.markdown('<div class="nav-header">🔍 QR Code Plate Processor</div>', unsafe_allow_html=True)
    
    # Check if QR libraries are available
    if not QR_AVAILABLE:
        st.error("❌ QR Code processing requires additional libraries!")
        
        # Show specific missing libraries
        missing_libs = []
        if not CV2_AVAILABLE:
            missing_libs.append("opencv-python")
        if not PYZBAR_AVAILABLE:
            missing_libs.append("pyzbar")
        
        st.markdown(f"""
        **Missing libraries:** {', '.join(missing_libs)}
        
        **For local development, install with:**
        ```bash
        pip install {' '.join(missing_libs)}
        ```
        
        **For Streamlit Cloud deployment:**
        
        Add this `packages.txt` file to your repository root:
        ```
        libzbar0
        ```
        
        And this `requirements.txt`:
        ```
        opencv-python-headless
        pyzbar
        ```
        
        Note: OpenCV can be challenging in cloud environments. Use `opencv-python-headless` for better compatibility.
        """)
        
        st.info("💡 **Alternative**: You can use the Unified Plant Data Processor which works without these dependencies.")
        return
    
    # Check if template files exist
    lamp_exists = check_template_exists(LAMP_TEMPLATE)
    qpcr_exists = check_template_exists(QPCR_TEMPLATE)
    
    if not lamp_exists and not qpcr_exists:
        st.error(f"❌ Template files not found in repository!")
        st.info(f"Please ensure '{LAMP_TEMPLATE}' and/or '{QPCR_TEMPLATE}' are in the same directory as this application.")
        return
    
    # Show available templates
    st.success("✅ Template files found:")
    col1, col2 = st.columns(2)
    with col1:
        if lamp_exists:
            st.success(f"✅ {LAMP_TEMPLATE}")
        else:
            st.warning(f"⚠️ {LAMP_TEMPLATE} not found")
    with col2:
        if qpcr_exists:
            st.success(f"✅ {QPCR_TEMPLATE}")
        else:
            st.warning(f"⚠️ {QPCR_TEMPLATE} not found")
    
    # Cleanup old files periodically
    cleanup_old_shares(24)
    
    # Step 1: Template selection
    st.header("🧪 Step 1: Select Template Type")
    template_options = []
    if lamp_exists:
        template_options.append("LAMP")
    if qpcr_exists:
        template_options.append("QPCR")
    
    if not template_options:
        st.error("No valid templates available.")
        return
    
    selected_template = st.radio(
        "Choose template for processing:",
        template_options,
        key="template_choice",
        help=f"Templates are loaded from the repository: {', '.join(template_options)}"
    )
    
    # Step 2: Image Processing Settings
    st.header("⚙️ Step 2: Image Processing Settings")
    
    # Fixed plate configuration (not user-configurable)
    plate_config = {
        "cols": 8,
        "rows": 12,
        "margin": 12,
        "crop_width": 2180,
        "crop_height": 3940
    }
    
    # Image scale factor (user-configurable within reasonable bounds)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        scale_factor = st.slider(
            "Image Scale Factor", 
            min_value=0.8, 
            max_value=2.5, 
            value=1.0, 
            step=0.1,
            key="scale_factor",
            help="Scale factor for image processing. Higher values (1.5-2.0) may improve QR detection but increase processing time."
        )
    
    plate_config["scale_factor"] = scale_factor
    
    if scale_factor != 1.0:
        st.info(f"🔍 Processing images at {scale_factor}x scale for enhanced QR detection")
    
    # Step 3: Upload images (with integrated file sharing)
    st.header("📷 Step 3: Upload Plate Images")
    
    # Add tabs for direct upload vs shared files
    upload_tab, share_tab = st.tabs(["📤 Direct Upload", "📱 Mobile Sharing"])
    
    uploaded_images = None
    
    with upload_tab:
        st.subheader("📤 Direct File Upload")
        
        # Show HEIC support status
        if HEIF_AVAILABLE:
            st.success("✅ HEIC/HEIF image format support is available")
        else:
            st.warning("⚠️ HEIC/HEIF support not available - please convert Apple HEIC files to JPG/PNG format first")
            st.info("To add HEIC support, install: `pip install pillow-heif`")
        
        uploaded_images = st.file_uploader(
            "Upload plate images",
            type=['jpg', 'jpeg', 'png', 'heic', 'heif'],
            accept_multiple_files=True,
            key="plate_images_direct",
            help="Upload laboratory plate images for QR code extraction. HEIC files from Apple devices are supported if pillow-heif is installed."
        )
    
    with share_tab:
        st.subheader("📱 Mobile → Computer Workflow")
        st.markdown("""
        **Perfect for:** Take pictures on phone → Access from computer → Rename → Process
        """)
        
        # Two sub-tabs: Upload for sharing and Access shared files
        share_upload_tab, share_access_tab = st.tabs(["📤 Upload from Mobile", "💻 Access from Computer"])
        
        with share_upload_tab:
            st.info("Upload files here (typically from mobile) to generate a share code for accessing from another device.")
            
            if HEIF_AVAILABLE:
                st.success("✅ HEIC/HEIF image format support is available")
            else:
                st.warning("⚠️ HEIC/HEIF support not available - please convert Apple HEIC files to JPG/PNG format first")
            
            mobile_uploaded_files = st.file_uploader(
                "Choose files to share",
                accept_multiple_files=True,
                type=['jpg', 'jpeg', 'png', 'heic', 'heif'],
                key="mobile_upload_qr",
                help="Upload images to share between devices"
            )
            
            if mobile_uploaded_files:
                st.subheader("📋 Files to Share")
                for file in mobile_uploaded_files:
                    file_size = len(file.getvalue()) / 1024  # KB
                    st.write(f"📄 {file.name} ({file_size:.1f} KB)")
                
                st.markdown('<div class="big-action-button share-button">', unsafe_allow_html=True)
                if st.button("🔗 Generate Share Code", key="generate_share_qr"):
                    try:
                        share_code = save_shared_files(mobile_uploaded_files)
                        st.success("✅ Files uploaded successfully!")
                        st.markdown(f'<div class="share-code">Share Code: {share_code}</div>', 
                                  unsafe_allow_html=True)
                        st.info(f"""
                        **How to use this code:**
                        1. Switch to "💻 Access from Computer" tab
                        2. Enter the code: **{share_code}**
                        3. Rename and download your files
                        4. Return to "📤 Direct Upload" tab to process
                        
                        ⏰ **Note**: This code expires in 24 hours
                        """)
                    except Exception as e:
                        st.error(f"❌ Error uploading files: {str(e)}")
                st.markdown('</div>', unsafe_allow_html=True)
        
        with share_access_tab:
            st.info("Enter a share code to access files uploaded from another device.")
            
            share_code_input = st.text_input(
                "Enter Share Code",
                placeholder="Enter 6-character code (e.g., ABC123)",
                max_chars=6,
                key="share_code_input_qr",
                help="Enter the 6-character code generated when uploading files"
            ).upper()
            
            if share_code_input and len(share_code_input) == 6:
                shared_data = load_shared_files(share_code_input)
                
                if shared_data:
                    files = shared_data["files"]
                    metadata = shared_data["metadata"]
                    
                    st.success(f"✅ Found {len(files)} shared files")
                    
                    upload_time = datetime.fromtimestamp(metadata.get("upload_time", 0))
                    st.write(f"**Uploaded:** {upload_time.strftime('%Y-%m-%d %H:%M:%S')}")
                    
                    # File renaming interface
                    st.subheader("🔧 Rename Files")
                    st.info("Rename files to more meaningful names before downloading.")
                    
                    # Initialize session state for file renaming
                    if 'shared_file_names_qr' not in st.session_state:
                        st.session_state.shared_file_names_qr = {}
                    
                    renamed_files = []
                    existing_names = set()
                    
                    for i, file in enumerate(files):
                        col1, col2 = st.columns([1, 2])
                        
                        with col1:
                            file_size = len(file.getvalue()) / 1024  # KB
                            st.write(f"📄 {file.name}")
                            st.write(f"Size: {file_size:.1f} KB")
                        
                        with col2:
                            # Generate safe default name if not already set
                            file_key = f"{share_code_input}_{file.name}"
                            if file_key not in st.session_state.shared_file_names_qr:
                                safe_name = generate_safe_filename(file.name, existing_names)
                                st.session_state.shared_file_names_qr[file_key] = safe_name
                            
                            new_name = st.text_input(
                                f"New name:",
                                value=st.session_state.shared_file_names_qr[file_key],
                                key=f"shared_rename_qr_{i}_{share_code_input}",
                                help="Enter a new filename (with extension)"
                            )
                            
                            # Update session state
                            if new_name != st.session_state.shared_file_names_qr[file_key]:
                                # Ensure uniqueness
                                safe_new_name = generate_safe_filename(new_name, existing_names)
                                st.session_state.shared_file_names_qr[file_key] = safe_new_name
                                if safe_new_name != new_name:
                                    st.warning(f"Name adjusted to avoid conflicts: {safe_new_name}")
                            
                            existing_names.add(st.session_state.shared_file_names_qr[file_key])
                            
                            # Create renamed file
                            renamed_file = io.BytesIO(file.getvalue())
                            renamed_file.name = st.session_state.shared_file_names_qr[file_key]
                            renamed_files.append(renamed_file)
                    
                    st.subheader("📥 Download Files")
                    st.info("After downloading, switch to 'Direct Upload' tab to process these renamed files.")
                    
                    # Individual file downloads
                    for i, (original_file, renamed_file) in enumerate(zip(files, renamed_files)):
                        col1, col2 = st.columns([3, 1])
                        
                        with col1:
                            if original_file.name != renamed_file.name:
                                st.write(f"📄 {original_file.name} → {renamed_file.name}")
                            else:
                                st.write(f"📄 {renamed_file.name}")
                        
                        with col2:
                            # Determine MIME type
                            if renamed_file.name.lower().endswith(('.jpg', '.jpeg')):
                                mime_type = 'image/jpeg'
                            elif renamed_file.name.lower().endswith('.png'):
                                mime_type = 'image/png'
                            else:
                                mime_type = 'application/octet-stream'
                            
                            st.download_button(
                                label="📥 Download",
                                data=renamed_file.getvalue(),
                                file_name=renamed_file.name,
                                mime=mime_type,
                                key=f"download_shared_qr_{i}_{share_code_input}"
                            )
                    
                    # Bulk download option
                    if len(files) > 1:
                        st.subheader("📦 Bulk Download")
                        import zipfile
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            for renamed_file in renamed_files:
                                zip_file.writestr(renamed_file.name, renamed_file.getvalue())
                        
                        zip_buffer.seek(0)
                        
                        col1, col2, col3 = st.columns([1, 2, 1])
                        with col2:
                            st.download_button(
                                label="📦 Download All Files (ZIP)",
                                data=zip_buffer.getvalue(),
                                file_name=f"shared_files_{share_code_input}.zip",
                                mime="application/zip",
                                key=f"bulk_download_shared_qr_{share_code_input}"
                            )
                else:
                    st.error("❌ Invalid share code or files have expired")
            elif share_code_input and len(share_code_input) != 6:
                st.warning("⚠️ Share code must be exactly 6 characters")
    
    if not uploaded_images:
        st.info("Please upload plate images to process using either the Direct Upload or Mobile Sharing tabs.")
        return
    
    # Process button
    st.header("🚀 Step 4: Process Images")
    st.markdown('<div class="big-action-button qr-button">', unsafe_allow_html=True)
    process_clicked = st.button("🔍 Process Plate Images", key="process_plates")
    st.markdown('</div>', unsafe_allow_html=True)
    
    if process_clicked:
        # Load the selected template
        template_file = LAMP_TEMPLATE if selected_template == "LAMP" else QPCR_TEMPLATE
        
        results = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, uploaded_image in enumerate(files_to_process):
            # Files from shared access already have their renamed names
            display_name = uploaded_image.name
            status_text.text(f"Processing {display_name}...")
            
            # Load fresh template buffer for each image
            template_buffer = load_template_from_file(template_file)
            if not template_buffer:
                st.error(f"❌ Failed to load template file: {template_file}")
                continue
            
            result, _, error = process_plate_image(uploaded_image, template_buffer, plate_config, plate_config.get("scale_factor", 1.0))
            
            if error:
                st.error(f"❌ Error processing {display_name}: {error}")
            elif result:
                base_name = os.path.splitext(display_name)[0]
                results.append({
                    'original_name': getattr(uploaded_image, 'name', f'shared_file_{i}'),
                    'display_name': display_name,
                    'base_name': base_name,
                    'result': result
                })
                st.success(f"✅ Successfully processed {display_name}")
            
            progress_bar.progress((i + 1) / len(files_to_process))
        
        status_text.text("Processing complete!")
        
        if results:
            # IMPROVED UI: Better layout for results
            st.header("📊 Processing Results")
            
            # Overall statistics in a more prominent display
            st.subheader("📈 Summary Statistics")
            
            total_plates = len(results)
            total_wells = sum(result['result']['total'] for result in results)
            total_success = sum(result['result']['success'] for result in results)
            total_failed = total_wells - total_success
            overall_success_rate = (total_success / total_wells * 100) if total_wells > 0 else 0
            
            # Better metrics display (5 columns instead of 4)
            metric_col1, metric_col2, metric_col3, metric_col4, metric_col5 = st.columns(5)
            with metric_col1:
                st.metric("Plates Processed", total_plates)
            with metric_col2:
                st.metric("Total Wells", total_wells)
            with metric_col3:
                st.metric("Successful Reads", total_success)
            with metric_col4:
                st.metric("Failed Reads", total_failed)
            with metric_col5:
                st.metric("Success Rate", f"{overall_success_rate:.1f}%")
            
            # Bulk download section
            st.header("📥 Download Results")
            
            # Create ZIP file for bulk download
            import zipfile
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for result_data in results:
                    result = result_data['result']
                    base_name = result_data['base_name']
                    excel_filename = f"{base_name}_{selected_template}_filled.xlsx"
                    
                    # Add Excel file to ZIP
                    zip_file.writestr(excel_filename, result['excel_buffer'].getvalue())
            
            zip_buffer.seek(0)
            
            # Bulk download button
            st.subheader("📦 Download All Files")
            download_col1, download_col2, download_col3 = st.columns([1, 2, 1])
            with download_col2:
                st.markdown('<div class="big-action-button download-button">', unsafe_allow_html=True)
                st.download_button(
                    label="📦 Download All Files (ZIP)",
                    data=zip_buffer.getvalue(),
                    file_name=f"QR_Processing_Results_{selected_template}.zip",
                    mime="application/zip",
                    key="bulk_download_qr"
                )
                st.markdown('</div>', unsafe_allow_html=True)
            
            st.markdown("---")
            
            # IMPROVED: Individual results with better layout
            st.subheader("📄 Individual File Results")
            
            for idx, result_data in enumerate(results):
                result = result_data['result']
                
                # Create a container for each file result
                with st.container():
                    st.markdown(f"### 📄 {result_data['display_name']}")
                    
                    # Create three columns for better layout
                    info_col, stats_col, download_col = st.columns([2, 2, 1])
                    
                    with info_col:
                        st.write(f"**Original Name:** {result_data['original_name']}")
                        st.write(f"**Template Used:** {selected_template}")
                        if result_data['original_name'] != result_data['display_name']:
                            st.write(f"**Renamed To:** {result_data['display_name']}")
                    
                    with stats_col:
                        success_rate = (result['success'] / result['total']) * 100 if result['total'] > 0 else 0
                        st.write(f"**Success Rate:** {success_rate:.1f}%")
                        st.write(f"**Wells Read:** {result['success']}/{result['total']}")
                        
                        if result['failed'] > 0:
                            failed_preview = ', '.join(result['failed_positions'][:5])
                            if len(result['failed_positions']) > 5:
                                failed_preview += f" (+{len(result['failed_positions']) - 5} more)"
                            st.write(f"**Failed Wells:** {failed_preview}")
                    
                    with download_col:
                        # Download Excel
                        excel_filename = f"{result_data['base_name']}_{selected_template}_filled.xlsx"
                        st.download_button(
                            label="📥 Download Excel",
                            data=result['excel_buffer'].getvalue(),
                            file_name=excel_filename,
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key=f"excel_{idx}_{result_data['base_name']}"
                        )
                    
                    # Show debug image in an expander below
                    with st.expander(f"🔍 View Annotated Image - {result_data['display_name']}", expanded=False):
                        # Create two columns for the image display
                        img_col1, img_col2 = st.columns([3, 1])
                        with img_col1:
                            st.image(result['debug_image'], caption=f"Processed plate with QR detection results")
                        with img_col2:
                            st.write("**Legend:**")
                            st.write("🟢 Green text: QR code detected")
                            st.write("🔴 Red text: No QR code found")
                            st.write("🔵 Blue rectangles: Processing grid")
                    
                    st.markdown("---")

def main():
    """Main application function."""
    # Sidebar navigation
    st.sidebar.title("The Riaz Machine")
    st.sidebar.markdown("---")
    
    # Navigation options (removed separate Mobile File Sharing)
    app_mode = st.sidebar.radio(
        "Choose Function:",
        [
            "🔄 Unified Plant Data Processor", 
            "🔍 QR Code Plate Processor"
        ],
        key="main_nav"
    )
    
    # Function descriptions in sidebar
    st.sidebar.markdown("---")
    with st.sidebar.expander("📖 Function Descriptions", expanded=False):
        st.markdown("""
        **🔄 Unified Plant Data Processor**
        - Combines Plant Data Processor and Headwaters Submission
        - Smart detection of special client formats
        - Fuzzy column mapping for various naming conventions
        - Multi-sheet processing when appropriate
        - High-performance vectorized operations
        
        **🔍 QR Code Plate Processor**
        - Process laboratory plate images
        - Extract QR codes automatically
        - Generate filled Excel templates
        - Support for LAMP and QPCR formats
        - **Integrated mobile file sharing**
        - File renaming capability
        - Improved orientation handling
        - Perfect mobile → computer workflow
        """)
    
    # Route to appropriate function
    if "Unified Plant Data Processor" in app_mode:
        unified_processor()
    elif "QR Code Plate Processor" in app_mode:
        qr_plate_processor()

if __name__ == "__main__":
    main()
